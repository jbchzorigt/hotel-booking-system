"""
End-to-end pytest suite for the hotel marketplace backend.

This is a single, ORDERED scenario expressed as ``test_*`` functions — not
independent unit tests. Later phases assert on ledger totals produced by
earlier phases, so the functions run in file-definition order and share a
session-scoped ``state`` bag, TestClient, and seeded database (see
``conftest.py``). Environment and DB owner credentials are env-driven with
local fallbacks; nothing is hardcoded to one machine.

Phases:
    A  seed + escrow capture idempotency        (async, app engine + Redis)
    B  manager CRUD, RLS, check-in + WS, checkout (sync, TestClient)
    D  marketplace search/book, restaurant, food orders
    E  platform-admin dashboards + xlsx export
    F  auth provisioning/login + police dashboard API
    C  ledger & police ground truth             (async, owner engine)
    G  janitor sweep                            (async, owner engine)
"""

from __future__ import annotations

import concurrent.futures
import json
import os
import uuid
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import Any

import pytest
import redis as sync_redis
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.database import get_engine
from app.core.redis import get_redis
from app.core.security import create_access_token
from app.models.domain import (
    Booking,
    BookingStatus,
    EscrowStatus,
    MinibarConsumption,
    PlatformAccount,
    PlatformLedgerEntry,
    PoliceMatch,
    Restaurant,
    Room,
    RoomType,
    SubscriptionPlan,
    Tenant,
    User,
    UserRole,
)
from app.services.payment_escrow_service import (
    EscrowService,
    InvalidEscrowStateError,
    PaymentMethod,
)
from app.services.police_service import compute_registry_hash
from tests.conftest import OWNER_URL, owner_session_ctx

REGISTRY = "УБ11111111"

# WebSocket receive-with-timeout: TestClient's ws.receive_text() blocks, so
# run it on a worker thread and time it out.
_ws_pool = concurrent.futures.ThreadPoolExecutor(max_workers=2)


def _ws_recv(ws, timeout: float) -> str:
    return _ws_pool.submit(ws.receive_text).result(timeout=timeout)


def _hdr(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _tok(role, tenant=None, restaurant=None, realm="app", sub=None) -> str:
    return create_access_token(
        subject=str(sub or uuid.uuid4()),
        role=role,
        realm=realm,
        tenant_id=tenant,
        restaurant_id=restaurant,
    )


# =========================================================================== #
# Fixtures: clean DB + seed, then role tokens
# =========================================================================== #
async def _truncate_and_seed(state: dict[str, Any]) -> None:
    engine = create_async_engine(OWNER_URL)
    now = datetime.now(timezone.utc)
    try:
        async with async_sessionmaker(engine, expire_on_commit=False)() as s:
            # Start from a clean slate so seeds (which assume empty tables)
            # are reproducible across local re-runs and CI.
            await s.execute(
                text(
                    "DO $$ DECLARE r RECORD; BEGIN "
                    "FOR r IN SELECT tablename FROM pg_tables "
                    "WHERE schemaname='public' AND tablename <> 'alembic_version' "
                    "LOOP EXECUTE 'TRUNCATE TABLE ' || quote_ident(r.tablename) "
                    "|| ' RESTART IDENTITY CASCADE'; END LOOP; END $$;"
                )
            )
            await s.commit()

            s.add(
                PlatformAccount(
                    currency="MNT",
                    balance=Decimal("0.00"),
                    commission_rate=Decimal("0.0500"),
                )
            )
            ta = Tenant(
                name="Blue Sky Hotel", slug="blue-sky", contact_email="a@a.mn",
                maps_lat=Decimal("47.918530"), maps_lng=Decimal("106.917701"),
                address="Peace Avenue 17, Ulaanbaatar",
                subscription_plan=SubscriptionPlan.MONTHS_12,
                subscription_started_at=now,
                subscription_expires_at=now + timedelta(days=365),
            )
            tb = Tenant(
                name="Steppe Inn", slug="steppe-inn", contact_email="b@b.mn",
                maps_lat=Decimal("47.900000"), maps_lng=Decimal("106.900000"),
                subscription_plan=SubscriptionPlan.MONTHS_3,
                subscription_started_at=now,
                subscription_expires_at=now + timedelta(days=90),
            )
            s.add_all([ta, tb])
            await s.flush()
            cleaner = User(
                email="cleaner@bluesky.mn", hashed_password="x",
                full_name="Cleaner One", role=UserRole.CLEANER, tenant_id=ta.id,
            )
            room = Room(
                tenant_id=ta.id, room_number="101", room_type=RoomType.DOUBLE,
                beds=2, floor=1, base_price=Decimal("100000.00"),
            )
            s.add_all([cleaner, room])
            await s.flush()
            booking = Booking(
                tenant_id=ta.id, room_id=room.id, code="BK-TEST01",
                guest_full_name="Unverified Alias", guest_phone="+976-99112233",
                check_in_date=date.today(),
                check_out_date=date.today() + timedelta(days=2),
                status=BookingStatus.CONFIRMED,
                nightly_rate=Decimal("100000.00"),
                total_amount=Decimal("200000.00"),
                commission_rate=Decimal("0.0500"),
                commission_amount=Decimal("0.00"),
            )
            s.add(booking)
            s.add(
                WantedPerson(
                    registry_hash=compute_registry_hash(REGISTRY),
                    full_name="Мөнх Болд", address="unknown",
                    case_reference="CASE-42",
                )
            )
            await s.commit()
            state.update(
                tenant_a=ta.id, tenant_b=tb.id, cleaner=cleaner.id,
                room=room.id, booking=booking.id,
            )
    finally:
        await engine.dispose()


# WantedPerson import kept local to avoid a wide models import line above.
from app.models.domain import WantedPerson  # noqa: E402


@pytest.fixture(scope="session")
def seeded_db(state: dict[str, Any]) -> dict[str, Any]:
    """Truncate + seed once per session (sync wrapper: fully self-contained
    loop, engine created and disposed inside)."""
    import asyncio

    sync_redis.Redis(
        host=os.environ.get("REDIS_HOST", "localhost"),
        port=int(os.environ.get("REDIS_PORT", "6379")),
    ).flushdb()
    asyncio.run(_truncate_and_seed(state))
    return state


@pytest.fixture(scope="session")
def tokens(seeded_db: dict[str, Any]) -> dict[str, str]:
    ta, tb = seeded_db["tenant_a"], seeded_db["tenant_b"]
    return {
        "mgr_a": _tok("MANAGER", tenant=ta),
        "mgr_b": _tok("MANAGER", tenant=tb),
        "rec_a": _tok("RECEPTION", tenant=ta),
        "rec_b": _tok("RECEPTION", tenant=tb),
        "cln_a": _tok("CLEANER", tenant=ta, sub=seeded_db["cleaner"]),
        "police": _tok("POLICE", realm="police", sub="dispatch-01"),
        "admin": _tok("PLATFORM_ADMIN"),
    }


# =========================================================================== #
# Phase A — escrow capture idempotency (async; app engine + Redis)
# =========================================================================== #
async def test_a_escrow_capture_idempotency(seeded_db: dict[str, Any]) -> None:
    escrow = EscrowService()
    booking_id = seeded_db["booking"]
    key = f"cap-{uuid.uuid4()}"
    r1 = await escrow.pay_booking(
        booking_id, method=PaymentMethod.QPAY, idempotency_key=key
    )
    assert r1.escrow_status == "HELD" and r1.amount == "200000.00"

    r2 = await escrow.pay_booking(
        booking_id, method=PaymentMethod.QPAY, idempotency_key=key
    )
    assert r2 == r1, "replay must return the identical cached receipt"

    with pytest.raises(InvalidEscrowStateError):
        await escrow.pay_booking(
            booking_id, method=PaymentMethod.QPAY,
            idempotency_key=f"cap-{uuid.uuid4()}",
        )

    # Release loop-bound app resources before the TestClient's portal loop
    # rebinds them (asyncpg/redis connections are event-loop bound).
    await get_engine().dispose()
    await get_redis().aclose()


# =========================================================================== #
# Phase B — API flows through the TestClient
# =========================================================================== #
def test_b_manager_crud(client, tokens, state) -> None:
    mgr_a = tokens["mgr_a"]
    r = client.post("/api/v1/manager/minibar/categories",
                    json={"name": "Beverages"}, headers=_hdr(mgr_a))
    assert r.status_code == 201, r.text
    state["cat_id"] = r.json()["id"]
    r = client.post("/api/v1/manager/minibar/items",
                    json={"category_id": state["cat_id"],
                          "name": "Chinggis Beer 0.5L", "price": "8000.00"},
                    headers=_hdr(mgr_a))
    assert r.status_code == 201, r.text
    state["item_id"] = r.json()["id"]
    r = client.post("/api/v1/manager/minibar/categories",
                    json={"name": "Beverages"}, headers=_hdr(mgr_a))
    assert r.status_code == 409, "duplicate category must 409"


def test_b_rls_isolation(client, tokens, state) -> None:
    mgr_b, rec_b = tokens["mgr_b"], tokens["rec_b"]
    assert client.get("/api/v1/manager/rooms", headers=_hdr(mgr_b)).json() == []
    assert client.get("/api/v1/manager/minibar/items",
                      headers=_hdr(mgr_b)).json() == []
    r = client.post("/api/v1/reception/check-in",
                    json={"booking_id": str(state["booking"]),
                          "registry_number": REGISTRY}, headers=_hdr(rec_b))
    assert r.status_code == 404, "hotel B must not even learn booking A exists"


def test_b_checkin_ws_alert_and_minibar(client, tokens, state) -> None:
    rec_a, cln_a, police = tokens["rec_a"], tokens["cln_a"], tokens["police"]
    with client.websocket_connect(f"/ws/police/alerts?token={police}") as ws_pol, \
         client.websocket_connect(f"/ws/reception?token={rec_a}") as ws_rec:
        import time
        time.sleep(0.3)

        r = client.post("/api/v1/reception/check-in",
                        json={"booking_id": str(state["booking"]),
                              "registry_number": REGISTRY}, headers=_hdr(rec_a))
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["status"] == "CHECKED_IN" and body["room_number"] == "101"
        assert body["verified_full_name"] and body["verified_address"]

        alert = json.loads(_ws_recv(ws_pol, 15))
        assert alert["type"] == "POLICE_MATCH_ALERT"
        assert alert["wanted_full_name"] == "Мөнх Болд"
        assert alert["case_reference"] == "CASE-42"
        assert alert["room_number"] == "101"
        assert alert["hotel_name"] == "Blue Sky Hotel"
        assert abs(alert["hotel_maps_lat"] - 47.918530) < 1e-6

        r = client.post("/api/v1/cleaner/minibar/report",
                        json={"room_id": str(state["room"]),
                              "items": [{"minibar_item_id": state["item_id"],
                                         "quantity": 2}]}, headers=_hdr(cln_a))
        assert r.status_code == 201, r.text
        assert Decimal(r.json()["total_amount"]) == Decimal("16000.00")
        note = json.loads(_ws_recv(ws_rec, 10))
        assert note["type"] == "MINIBAR_REPORT" and note["room_number"] == "101"
        assert Decimal(note["total_amount"]) == Decimal("16000.00")


def test_b_checkout_and_housekeeping(client, tokens, state) -> None:
    rec_a, cln_a = tokens["rec_a"], tokens["cln_a"]
    assert client.get("/api/v1/cleaner/rooms/dirty",
                      headers=_hdr(cln_a)).json() == []

    r = client.post("/api/v1/reception/checkout",
                    json={"booking_id": str(state["booking"])}, headers=_hdr(rec_a))
    assert r.status_code == 200, r.text
    out = r.json()
    assert Decimal(str(out["total_amount"])) == Decimal("200000.00")
    assert Decimal(str(out["commission_amount"])) == Decimal("10000.00")
    assert Decimal(str(out["hotel_amount"])) == Decimal("190000.00")
    assert Decimal(str(out["minibar_charged"])) == Decimal("16000.00")
    assert out["room_state"] == "VACANT_DIRTY"

    r = client.post("/api/v1/reception/checkout",
                    json={"booking_id": str(state["booking"])}, headers=_hdr(rec_a))
    assert r.status_code == 409, "second checkout must be rejected"

    dirty = client.get("/api/v1/cleaner/rooms/dirty", headers=_hdr(cln_a)).json()
    assert [d["room_number"] for d in dirty] == ["101"]
    assert "base_price" not in dirty[0] and "guest" not in json.dumps(dirty)
    r = client.post(f"/api/v1/cleaner/rooms/{state['room']}/mark-clean",
                    headers=_hdr(cln_a))
    assert r.status_code == 200 and r.json()["state"] == "VACANT_CLEAN"


# =========================================================================== #
# Phase D — marketplace, restaurant, food orders
# =========================================================================== #
def test_d_marketplace_search_and_booking(client, state) -> None:
    ci = (date.today() + timedelta(days=3)).isoformat()
    co = (date.today() + timedelta(days=5)).isoformat()
    state["ci"], state["co"] = ci, co

    r = client.get("/api/v1/marketplace/search",
                   params={"lat": 47.918, "lng": 106.917, "radius_km": 5,
                           "check_in": ci, "check_out": co})
    assert r.status_code == 200, r.text
    hotels = {h["slug"]: h for h in r.json()}
    assert "blue-sky" in hotels and "steppe-inn" in hotels
    assert hotels["blue-sky"]["available_rooms"] == 1
    assert Decimal(str(hotels["blue-sky"]["min_nightly_rate"])) == Decimal("100000.00")
    assert hotels["blue-sky"]["distance_km"] < 1.0
    assert hotels["steppe-inn"]["available_rooms"] == 0

    payload = {"room_id": str(state["room"]), "guest_full_name": "Second Guest",
               "guest_phone": "+976-88112233", "check_in_date": ci,
               "check_out_date": co, "payment_method": "QPAY"}
    r = client.post("/api/v1/marketplace/book", json=payload,
                    headers={"Idempotency-Key": f"book-{uuid.uuid4()}"})
    assert r.status_code == 201, r.text
    b2 = r.json()
    assert b2["status"] == "CONFIRMED" and b2["escrow_status"] == "HELD"
    assert Decimal(str(b2["total_amount"])) == Decimal("200000.00")
    state["b2"] = b2

    r = client.post("/api/v1/marketplace/book", json=payload,
                    headers={"Idempotency-Key": f"book-{uuid.uuid4()}"})
    assert r.status_code == 409, "GiST exclusion must reject the overlap"
    r = client.get("/api/v1/marketplace/search",
                   params={"lat": 47.918, "lng": 106.917, "radius_km": 5,
                           "check_in": ci, "check_out": co})
    assert {h["slug"]: h for h in r.json()}["blue-sky"]["available_rooms"] == 0


def test_d_restaurant_menu(client, tokens, state) -> None:
    mgr_a = tokens["mgr_a"]
    r = client.post("/api/v1/manager/restaurants",
                    json={"name": "Modern Nomads", "phone": "+976-70110011"},
                    headers=_hdr(mgr_a))
    assert r.status_code == 201, r.text
    state["rest_id"] = r.json()["id"]
    owner = _tok("RESTAURANT_OWNER", restaurant=uuid.UUID(state["rest_id"]))
    state["owner_token"] = owner

    r = client.post("/api/v1/restaurant/menu-items",
                    json={"name": "Khuushuur", "category": "Mains",
                          "price": "5000.00"}, headers=_hdr(owner))
    assert r.status_code == 201, r.text
    state["food_id"] = r.json()["id"]
    assert client.patch(f"/api/v1/restaurant/menu-items/{state['food_id']}",
                        json={"description": "Fried meat pastry"},
                        headers=_hdr(owner)).status_code == 200
    stranger = _tok("RESTAURANT_OWNER", restaurant=uuid.uuid4())
    assert client.get("/api/v1/restaurant/menu-items",
                      headers=_hdr(stranger)).json() == []
    assert client.patch(f"/api/v1/restaurant/menu-items/{state['food_id']}",
                        json={"price": "1.00"},
                        headers=_hdr(stranger)).status_code == 404
    r = client.get(f"/api/v1/marketplace/restaurants/{state['rest_id']}/menu")
    assert r.status_code == 200 and r.json()["items"][0]["name"] == "Khuushuur"


def test_d_food_order(client, tokens, state) -> None:
    rec_a = tokens["rec_a"]
    owner = state["owner_token"]
    r = client.post("/api/v1/reception/check-in",
                    json={"booking_id": state["b2"]["booking_id"],
                          "registry_number": "АА22222222"}, headers=_hdr(rec_a))
    assert r.status_code == 200, r.text

    with client.websocket_connect(f"/ws/restaurant/orders?token={owner}") as ws_own:
        import time
        time.sleep(0.3)
        r = client.post("/api/v1/marketplace/order",
                        json={"booking_code": state["b2"]["booking_code"],
                              "restaurant_id": state["rest_id"],
                              "items": [{"food_item_id": state["food_id"],
                                         "quantity": 3}]},
                        headers={"Idempotency-Key": f"food-{uuid.uuid4()}"})
        assert r.status_code == 201, r.text
        o = r.json()
        assert Decimal(str(o["total_amount"])) == Decimal("15000.00")
        assert o["escrow_status"] == "HELD" and o["room_number"] == "101"
        state["oid"] = o["order_id"]
        kitchen = json.loads(_ws_recv(ws_own, 10))
        assert kitchen["type"] == "NEW_FOOD_ORDER"
        assert kitchen["room_number"] == "101"
        assert Decimal(kitchen["total_amount"]) == Decimal("15000.00")


def test_d_food_fulfilment(client, state) -> None:
    owner, oid = state["owner_token"], state["oid"]
    for st in ("ACCEPTED", "PREPARING", "DELIVERED"):
        r = client.patch(f"/api/v1/restaurant/orders/{oid}/status",
                         json={"status": st}, headers=_hdr(owner))
        assert r.status_code == 200, r.text
    assert r.json()["escrow_status"] == "RELEASED"
    assert client.patch(f"/api/v1/restaurant/orders/{oid}/status",
                        json={"status": "ACCEPTED"},
                        headers=_hdr(owner)).status_code == 409


# =========================================================================== #
# Phase E — platform-admin dashboards + xlsx export
# =========================================================================== #
def test_e_admin_dashboards(client, tokens) -> None:
    mgr_a, admin = tokens["mgr_a"], tokens["admin"]
    assert client.get("/api/v1/admin/dashboard/revenue",
                      headers=_hdr(mgr_a)).status_code == 403
    r = client.get("/api/v1/admin/dashboard/revenue", headers=_hdr(admin))
    assert r.status_code == 200, r.text
    d = r.json()
    assert Decimal(str(d["wallet_balance"])) == Decimal("11550.00")
    assert Decimal(str(d["total_commission_collected"])) == Decimal("11550.00")
    assert set(d["by_source"]) == {"BOOKING_COMMISSION", "MINIBAR_COMMISSION",
                                   "FOOD_ORDER_COMMISSION"}

    top = client.get("/api/v1/admin/dashboard/top-rooms",
                     headers=_hdr(admin)).json()
    assert top[0]["room_number"] == "101" and top[0]["demand"] == 2
    assert Decimal(str(top[0]["gross_revenue"])) == Decimal("400000.00")

    r = client.get("/api/v1/admin/export/revenue", headers=_hdr(admin))
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in r.headers["content-disposition"]
    import io as _io

    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(r.content))
    assert wb.sheetnames == ["Monthly Revenue", "Minibar Statistics"]
    ws1 = wb["Monthly Revenue"]
    sources = {ws1.cell(row=i, column=2).value for i in range(2, ws1.max_row)}
    assert {"BOOKING_COMMISSION", "MINIBAR_COMMISSION",
            "FOOD_ORDER_COMMISSION"} <= sources
    ws2 = wb["Minibar Statistics"]
    assert ws2.cell(row=2, column=2).value == "Chinggis Beer 0.5L"
    assert ws2.cell(row=2, column=3).value == 2


# =========================================================================== #
# Phase F — auth provisioning/login + police dashboard API
# =========================================================================== #
def test_f_auth_and_police(client, tokens) -> None:
    mgr_a, police = tokens["mgr_a"], tokens["police"]
    r = client.post("/api/v1/auth/users",
                    json={"email": "Reception2@BlueSky.mn",
                          "password": "S3cure-Front-Desk!",
                          "full_name": "Front Desk Two", "role": "RECEPTION"},
                    headers=_hdr(mgr_a))
    assert r.status_code == 201, r.text
    assert r.json()["email"] == "reception2@bluesky.mn"
    r = client.post("/api/v1/auth/users",
                    json={"email": "reception2@bluesky.mn",
                          "password": "S3cure-Front-Desk!",
                          "full_name": "Dup", "role": "RECEPTION"},
                    headers=_hdr(mgr_a))
    assert r.status_code == 409, "duplicate email must 409"
    r = client.post("/api/v1/auth/users",
                    json={"email": "evil@x.mn", "password": "Whatever123!",
                          "full_name": "Evil", "role": "PLATFORM_ADMIN"},
                    headers=_hdr(mgr_a))
    assert r.status_code == 403, "manager must not mint platform admins"

    assert client.post("/api/v1/auth/login",
                       json={"email": "reception2@bluesky.mn",
                             "password": "wrong-password-1"}).status_code == 401
    assert client.post("/api/v1/auth/login",
                       json={"email": "ghost@nowhere.mn",
                             "password": "wrong-password-1"}).status_code == 401
    r = client.post("/api/v1/auth/login",
                    json={"email": "reception2@bluesky.mn",
                          "password": "S3cure-Front-Desk!"})
    assert r.status_code == 200, r.text
    login = r.json()
    assert login["role"] == "RECEPTION" and login["token_type"] == "bearer"
    r = client.post("/api/v1/reception/checkout",
                    json={"booking_id": str(uuid.uuid4())},
                    headers=_hdr(login["access_token"]))
    assert r.status_code == 404, r.text

    assert client.get("/api/v1/police/matches",
                      headers=_hdr(mgr_a)).status_code == 403
    r = client.get("/api/v1/police/matches", headers=_hdr(police))
    assert r.status_code == 200, r.text
    matches = r.json()
    assert len(matches) == 1
    m = matches[0]
    assert m["wanted_full_name"] == "Мөнх Болд" and m["room_number"] == "101"
    assert m["hotel_name"] == "Blue Sky Hotel" and m["status"] == "PENDING_REVIEW"
    r = client.post(f"/api/v1/police/matches/{m['match_id']}/resolve",
                    json={"resolution": "CONFIRMED", "note": "Unit dispatched"},
                    headers=_hdr(police))
    assert r.status_code == 200 and r.json()["status"] == "CONFIRMED"
    r = client.post(f"/api/v1/police/matches/{m['match_id']}/resolve",
                    json={"resolution": "DISMISSED"}, headers=_hdr(police))
    assert r.status_code == 409, "double resolution must 409"


# =========================================================================== #
# Phase C — ledger & police ground truth (async; owner engine)
# =========================================================================== #
async def test_c_ledger_and_police_truth(state) -> None:
    async with owner_session_ctx() as s:
        platform = (await s.execute(select(PlatformAccount))).scalar_one()
        assert platform.balance == Decimal("11550.00"), platform.balance
        hotel = await s.get(Tenant, state["tenant_a"])
        assert hotel.wallet_balance == Decimal("205200.00"), hotel.wallet_balance
        resto = (await s.execute(select(Restaurant))).scalar_one()
        assert resto.wallet_balance == Decimal("14250.00"), resto.wallet_balance
        entries = (
            await s.execute(
                select(PlatformLedgerEntry).order_by(PlatformLedgerEntry.created_at)
            )
        ).scalars().all()
        assert [e.source_type.value for e in entries] == [
            "MINIBAR_COMMISSION", "BOOKING_COMMISSION", "FOOD_ORDER_COMMISSION"
        ]
        assert entries[-1].balance_after == Decimal("11550.00")

        booking = await s.get(Booking, state["booking"])
        assert booking.status == BookingStatus.CHECKED_OUT
        assert booking.escrow_status == EscrowStatus.RELEASED
        assert booking.commission_amount == Decimal("10000.00")
        assert booking.guest_registry_hash == compute_registry_hash(REGISTRY)
        bookings_total = (
            await s.execute(select(func.count()).select_from(Booking))
        ).scalar_one()
        assert bookings_total == 2
        match = (await s.execute(select(PoliceMatch))).scalar_one()
        assert match.booking_id == booking.id
        assert match.status.value == "CONFIRMED"
        assert match.review_note == "Unit dispatched"
        settled = (
            await s.execute(
                select(func.count()).select_from(MinibarConsumption)
                .where(MinibarConsumption.is_settled)
            )
        ).scalar_one()
        assert settled == 1


# =========================================================================== #
# Phase G — janitor sweep (async; owner engine, no app-global engine)
# =========================================================================== #
async def test_g_janitor_sweep(state) -> None:
    from app.services.janitor_service import JanitorService

    engine = create_async_engine(OWNER_URL)

    @asynccontextmanager
    async def owner_scope():
        async with async_sessionmaker(engine, expire_on_commit=False)() as s:
            async with s.begin():
                yield s

    def _pending(code, days_in, days_out, total):
        return Booking(
            tenant_id=state["tenant_a"], room_id=state["room"], code=code,
            guest_full_name="Cart Guest", guest_phone="+976-00000000",
            check_in_date=date.today() + timedelta(days=days_in),
            check_out_date=date.today() + timedelta(days=days_out),
            status=BookingStatus.PENDING,
            nightly_rate=Decimal("100000.00"), total_amount=Decimal(total),
            commission_rate=Decimal("0.0500"), commission_amount=Decimal("0.00"),
        )

    try:
        async with async_sessionmaker(engine, expire_on_commit=False)() as s:
            stale = _pending("BK-STALE1", 20, 22, "200000.00")
            fresh = _pending("BK-FRESH1", 25, 26, "100000.00")
            s.add_all([stale, fresh])
            await s.flush()
            await s.execute(text(
                "UPDATE bookings SET created_at = now() - interval '20 minutes' "
                "WHERE code = 'BK-STALE1'"))
            await s.commit()
            stale_id, fresh_id = stale.id, fresh.id

        result = await JanitorService(session_scope=owner_scope).sweep_once()
        assert result["skipped"] is False
        assert result["bookings_cancelled"] == 1, result

        async with async_sessionmaker(engine, expire_on_commit=False)() as s:
            assert (await s.get(Booking, stale_id)).status == BookingStatus.CANCELLED
            assert (await s.get(Booking, fresh_id)).status == BookingStatus.PENDING
            # The GiST slot is freed: rebooking the stale dates now succeeds.
            s.add(_pending("BK-RETRY1", 20, 22, "200000.00"))
            await s.commit()
    finally:
        await engine.dispose()


# =========================================================================== #
# Phase H — B2C in-room dining: provisioning, RLS, QPay-invoiced orders
# =========================================================================== #
def test_h_provision_restaurant_with_manager(client, tokens, state) -> None:
    mgr_a = tokens["mgr_a"]
    r = client.post("/api/v1/manager/restaurants/provision", headers=_hdr(mgr_a),
                    json={"name": "Khan Buuz", "phone": "+976-70123456",
                          "manager_email": "Manager@KhanBuuz.mn",
                          "manager_password": "Khan-Buuz-2026!",
                          "manager_full_name": "Buuz Manager"})
    assert r.status_code == 201, r.text
    prov = r.json()
    assert prov["manager_role"] == "RESTAURANT_OWNER"
    assert prov["manager_email"] == "manager@khanbuuz.mn"   # normalised
    state["rest2_id"] = prov["restaurant_id"]

    # Duplicate restaurant name -> 409
    r = client.post("/api/v1/manager/restaurants/provision", headers=_hdr(mgr_a),
                    json={"name": "Khan Buuz", "manager_email": "x@y.mn",
                          "manager_password": "Whatever-123!",
                          "manager_full_name": "Dup"})
    assert r.status_code == 409, r.text

    # ATOMICITY: new name but duplicate email -> 409 AND the restaurant
    # must NOT have been half-created (both rows or neither).
    r = client.post("/api/v1/manager/restaurants/provision", headers=_hdr(mgr_a),
                    json={"name": "Ghost Diner",
                          "manager_email": "manager@khanbuuz.mn",
                          "manager_password": "Whatever-123!",
                          "manager_full_name": "Dup"})
    assert r.status_code == 409, r.text
    names = [x["name"] for x in client.get("/api/v1/manager/restaurants",
                                           headers=_hdr(mgr_a)).json()]
    assert "Ghost Diner" not in names, "half-created restaurant leaked"

    # The provisioned manager can REALLY log in (bcrypt roundtrip)
    r = client.post("/api/v1/auth/login",
                    json={"email": "manager@khanbuuz.mn",
                          "password": "Khan-Buuz-2026!"})
    assert r.status_code == 200, r.text
    assert r.json()["role"] == "RESTAURANT_OWNER"
    state["mgr2_token"] = r.json()["access_token"]


def test_h_restaurant_manager_rls_boundaries(client, state) -> None:
    m2 = state["mgr2_token"]
    # Own realm: create + see ONLY own menu
    r = client.post("/api/v1/restaurant/menu-items", headers=_hdr(m2),
                    json={"name": "Buuz (10 pcs)", "category": "Mains",
                          "price": "12000.00"})
    assert r.status_code == 201, r.text
    state["buuz_id"] = r.json()["id"]
    mine = client.get("/api/v1/restaurant/menu-items", headers=_hdr(m2)).json()
    assert [i["name"] for i in mine] == ["Buuz (10 pcs)"], (
        "manager must not see Modern Nomads' menu")
    # Cannot touch the OTHER restaurant's item
    assert client.patch(f"/api/v1/restaurant/menu-items/{state['food_id']}",
                        json={"price": "1.00"},
                        headers=_hdr(m2)).status_code == 404
    # NO access to hotel / booking / admin surfaces
    assert client.get("/api/v1/manager/rooms", headers=_hdr(m2)).status_code == 403
    assert client.get("/api/v1/reception/bookings",
                      headers=_hdr(m2)).status_code == 403
    assert client.get("/api/v1/admin/dashboard/revenue",
                      headers=_hdr(m2)).status_code == 403
    assert client.get("/api/v1/police/matches", headers=_hdr(m2)).status_code == 403


def test_h_public_booking_restaurants(client, state) -> None:
    b2_id = state["b2"]["booking_id"]
    r = client.get(f"/api/v1/public/bookings/{b2_id}/restaurants")
    assert r.status_code == 200, r.text
    listing = {x["name"]: x for x in r.json()}
    assert "Khan Buuz" in listing and "Modern Nomads" in listing
    assert [i["name"] for i in listing["Khan Buuz"]["items"]] == ["Buuz (10 pcs)"]
    # Unknown booking -> 404; terminated stay (guest1, CHECKED_OUT) -> 409
    assert client.get(
        f"/api/v1/public/bookings/{uuid.uuid4()}/restaurants").status_code == 404
    assert client.get(
        f"/api/v1/public/bookings/{state['booking']}/restaurants"
    ).status_code == 409


def test_h_public_order_and_concurrent_webhook(client, state) -> None:
    from app.services.qpay_service import sign_webhook

    b2_id = state["b2"]["booking_id"]
    # Unavailable/foreign item -> 422, nothing written
    r = client.post(f"/api/v1/public/bookings/{b2_id}/orders",
                    json={"restaurant_id": state["rest2_id"],
                          "items": [{"menu_item_id": str(uuid.uuid4()),
                                     "quantity": 1}]})
    assert r.status_code == 422, r.text

    # Create the UNPAID order + QPay invoice
    r = client.post(f"/api/v1/public/bookings/{b2_id}/orders",
                    json={"restaurant_id": state["rest2_id"],
                          "items": [{"menu_item_id": state["buuz_id"],
                                     "quantity": 2}]})
    assert r.status_code == 201, r.text
    order = r.json()
    assert order["status"] == "PLACED"
    assert order["escrow_status"] == "NOT_FUNDED"
    assert Decimal(str(order["total_amount"])) == Decimal("24000.00")
    invoice_id = order["qpay_invoice"]["invoice_id"]
    state["order2_id"] = order["order_id"]

    # Restaurant feed hides UNPAID orders
    m2 = state["mgr2_token"]
    feed = client.get("/api/v1/restaurant/orders", headers=_hdr(m2)).json()
    assert all(o["id"] != order["order_id"] for o in feed), (
        "kitchen must not see unpaid orders")

    # Fire the PAID webhook 5x CONCURRENTLY — exactly one funds the order.
    body = json.dumps({"invoice_id": invoice_id,
                       "payment_status": "PAID"}).encode()
    hdrs = {"X-QPay-Signature": sign_webhook(body),
            "Content-Type": "application/json"}
    with client.websocket_connect(
            f"/ws/restaurant/orders?token={m2}") as ws_kitchen:
        import time
        time.sleep(0.3)
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as pool5:
            results = list(pool5.map(
                lambda _: client.post("/api/v1/payments/qpay-webhook",
                                      content=body, headers=hdrs),
                range(5)))
        assert all(r.status_code == 200 for r in results)
        outcomes = [r.json()["result"] for r in results]
        assert outcomes.count("funded") == 1, outcomes
        assert outcomes.count("already_funded") == 4, outcomes
        funded = next(r.json() for r in results if r.json()["result"] == "funded")
        assert funded["kind"] == "food_order"

        # Money first, kitchen second: WS alert arrives AFTER funding
        kitchen = json.loads(_ws_recv(ws_kitchen, 10))
        assert kitchen["type"] == "NEW_FOOD_ORDER"
        assert kitchen["order_id"] == order["order_id"]
        assert Decimal(kitchen["total_amount"]) == Decimal("24000.00")

    # Bad signature is still rejected
    assert client.post("/api/v1/payments/qpay-webhook", content=body,
                       headers={"X-QPay-Signature": "deadbeef",
                                "Content-Type": "application/json"}
                       ).status_code == 401


def test_h_order_funded_visibility(client, state) -> None:
    # Guest poll flips to funded
    r = client.get(f"/api/v1/public/orders/{state['order2_id']}")
    assert r.status_code == 200, r.text
    poll = r.json()
    assert poll["is_funded"] is True and poll["escrow_status"] == "HELD"
    assert poll["status"] == "PLACED" and poll["paid_at"] is not None

    # Kitchen now sees the paid order and can fulfil it end-to-end
    m2 = state["mgr2_token"]
    feed = client.get("/api/v1/restaurant/orders", headers=_hdr(m2)).json()
    mine = [o for o in feed if o["id"] == state["order2_id"]]
    assert len(mine) == 1 and mine[0]["escrow_status"] == "HELD"
    for st in ("ACCEPTED", "PREPARING", "DELIVERED"):
        r = client.patch(
            f"/api/v1/restaurant/orders/{state['order2_id']}/status",
            json={"status": st}, headers=_hdr(m2))
        assert r.status_code == 200, r.text
    assert r.json()["escrow_status"] == "RELEASED"  # 95% -> restaurant wallet


def test_h_attach_manager_to_existing_restaurant(client, tokens, state) -> None:
    """POST /restaurants/{id}/manager — credentials for a restaurant that
    was registered WITHOUT a manager (Modern Nomads, from Phase D)."""
    admin_a = _tok("HOTEL_ADMIN", tenant=state["tenant_a"])
    admin_b = _tok("HOTEL_ADMIN", tenant=state["tenant_b"])
    url = f"/api/v1/restaurants/{state['rest_id']}/manager"
    payload = {"email": "Chef@ModernNomads.mn", "password": "Nomads-Chef-26!",
               "full_name": "Nomads Chef"}

    # Security matrix: only the OWNING hotel's HOTEL_ADMIN passes.
    assert client.post(url, json=payload).status_code == 401          # anon
    assert client.post(url, json=payload,
                       headers=_hdr(tokens["mgr_a"])).status_code == 403  # MANAGER
    assert client.post(url, json=payload,
                       headers=_hdr(tokens["rec_a"])).status_code == 403  # RECEPTION
    assert client.post(url, json=payload,
                       headers=_hdr(admin_b)).status_code == 404  # foreign hotel
    assert client.post(f"/api/v1/restaurants/{uuid.uuid4()}/manager",
                       json=payload, headers=_hdr(admin_a)).status_code == 404

    r = client.post(url, json=payload, headers=_hdr(admin_a))
    assert r.status_code == 201, r.text
    created = r.json()
    assert created["role"] == "RESTAURANT_OWNER"
    assert created["email"] == "chef@modernnomads.mn"          # normalised
    assert created["restaurant_id"] == state["rest_id"]
    assert created["tenant_id"] == str(state["tenant_a"])      # derived link
    # Duplicate email -> 409
    assert client.post(url, json=payload,
                       headers=_hdr(admin_a)).status_code == 409

    # The credentials WORK: login -> confined to Modern Nomads' realm.
    r = client.post("/api/v1/auth/login",
                    json={"email": "chef@modernnomads.mn",
                          "password": "Nomads-Chef-26!"})
    assert r.status_code == 200, r.text
    chef = r.json()["access_token"]
    menu = client.get("/api/v1/restaurant/menu-items", headers=_hdr(chef)).json()
    assert [i["name"] for i in menu] == ["Khuushuur"], (
        "chef must see Modern Nomads' menu and nothing else")
    # Not Khan Buuz's item, not hotel surfaces.
    assert client.patch(f"/api/v1/restaurant/menu-items/{state['buuz_id']}",
                        json={"price": "1.00"},
                        headers=_hdr(chef)).status_code == 404
    assert client.get("/api/v1/manager/rooms",
                      headers=_hdr(chef)).status_code == 403
    assert client.get("/api/v1/reception/bookings",
                      headers=_hdr(chef)).status_code == 403


def test_h_upload_and_menu_image(client, tokens, state) -> None:
    """POST /api/v1/upload: auth + content validation + static serving,
    then the URL flows onto a menu item and out through the public menu."""
    m2 = state["mgr2_token"]
    png = (b"\x89PNG\r\n\x1a\n" + b"\x00" * 64)  # valid magic, tiny body

    # Anonymous -> 401; guest-less roles that manage no content -> 403
    assert client.post("/api/v1/upload",
                       files={"file": ("a.png", png, "image/png")}
                       ).status_code == 401
    assert client.post("/api/v1/upload",
                       files={"file": ("a.png", png, "image/png")},
                       headers=_hdr(tokens["cln_a"])).status_code == 403
    # Wrong declared type -> 415; right type but fake content -> 415
    assert client.post("/api/v1/upload",
                       files={"file": ("x.pdf", b"%PDF-1.4", "application/pdf")},
                       headers=_hdr(m2)).status_code == 415
    assert client.post("/api/v1/upload",
                       files={"file": ("fake.png", b"<?php evil ?>", "image/png")},
                       headers=_hdr(m2)).status_code == 415

    # Happy path: traversal-attempt filename is discarded server-side
    r = client.post("/api/v1/upload",
                    files={"file": ("../../evil.png", png, "image/png")},
                    headers=_hdr(m2))
    assert r.status_code == 201, r.text
    up = r.json()
    assert up["url"].startswith("/static/uploads/")
    assert ".." not in up["url"] and up["filename"].endswith(".png")
    assert up["size_bytes"] == len(png)

    # The file is genuinely served back via StaticFiles
    served = client.get(up["url"])
    assert served.status_code == 200 and served.content == png

    # image_url flows onto a menu item and out through BOTH menus
    r = client.post("/api/v1/restaurant/menu-items", headers=_hdr(m2),
                    json={"name": "Tsuivan", "category": "Mains",
                          "price": "15000.00", "image_url": up["url"]})
    assert r.status_code == 201, r.text
    assert r.json()["image_url"] == up["url"]
    b2_id = state["b2"]["booking_id"]
    listing = {x["name"]: x for x in client.get(
        f"/api/v1/public/bookings/{b2_id}/restaurants").json()}
    tsuivan = next(i for i in listing["Khan Buuz"]["items"]
                   if i["name"] == "Tsuivan")
    assert tsuivan["image_url"] == up["url"]


def test_h_restaurant_has_manager_flag(client, tokens, state) -> None:
    """GET /manager/restaurants exposes onboarding status per restaurant."""
    mgr_a = tokens["mgr_a"]
    # A fresh restaurant with NO credentials yet
    r = client.post("/api/v1/manager/restaurants",
                    json={"name": "Unstaffed Cafe"}, headers=_hdr(mgr_a))
    assert r.status_code == 201, r.text

    flags = {x["name"]: x["has_manager"] for x in client.get(
        "/api/v1/manager/restaurants", headers=_hdr(mgr_a)).json()}
    assert flags["Khan Buuz"] is True         # provisioned with manager
    assert flags["Modern Nomads"] is True     # chef attached in prior test
    assert flags["Unstaffed Cafe"] is False   # nobody attached yet

    # Hotel B sees none of hotel A's restaurants at all (RLS unchanged)
    assert client.get("/api/v1/manager/restaurants",
                      headers=_hdr(tokens["mgr_b"])).json() == []
