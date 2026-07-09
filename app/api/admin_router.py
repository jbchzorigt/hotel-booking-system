"""
Platform-admin endpoints: revenue dashboard, demand analytics, Excel export.

All endpoints are gated to PLATFORM_ADMIN; ``get_scoped_session`` injects
``platform_session()``, whose RLS identity is the only one allowed to read
``platform_accounts`` / ``platform_ledger_entries`` — and which sees every
tenant, because reconciliation is legitimately cross-tenant.

Reporting reads the LEDGER, not the wallet: ``PlatformAccount.balance`` is
a cache for hot-path credit/debit; every aggregate below is derived from
the append-only ``platform_ledger_entries`` so the numbers are auditable.
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import Annotated

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import case, func, select

from app.dependencies.auth import AuthContext, ScopedSession, require_roles
from app.models.domain import (
    Booking,
    BookingStatus,
    LedgerDirection,
    MinibarConsumption,
    PlatformAccount,
    PlatformLedgerEntry,
    Room,
    Tenant,
    UserRole,
)

router = APIRouter(prefix="/admin", tags=["platform-admin"])

AdminCtx = Annotated[AuthContext, Depends(require_roles(UserRole.PLATFORM_ADMIN))]

_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# ===========================================================================
# Schemas
# ===========================================================================
class RevenueDashboard(BaseModel):
    currency: str
    #: Live wallet balance (cache; reconciles against the ledger below).
    wallet_balance: Decimal
    commission_rate: Decimal
    #: Lifetime commission credited, from the ledger (auditable truth).
    total_commission_collected: Decimal
    #: Lifetime debits (payouts, refunds).
    total_debited: Decimal
    #: Credit breakdown by source (BOOKING_COMMISSION, MINIBAR_COMMISSION, …).
    by_source: dict[str, Decimal]
    ledger_entries: int


class TopRoom(BaseModel):
    room_id: uuid.UUID
    room_number: str
    hotel_name: str
    #: Bookings that held or moved money (CANCELLED/NO_SHOW excluded).
    demand: int
    gross_revenue: Decimal


# ===========================================================================
# GET /dashboard/revenue
# ===========================================================================
@router.get("/dashboard/revenue", response_model=RevenueDashboard)
async def revenue_dashboard(
    ctx: AdminCtx, session: ScopedSession
) -> RevenueDashboard:
    """Platform commission position: wallet + ledger-derived aggregates."""
    account = (
        await session.execute(select(PlatformAccount).limit(1))
    ).scalar_one()

    rows = (
        await session.execute(
            select(
                PlatformLedgerEntry.direction,
                PlatformLedgerEntry.source_type,
                func.coalesce(func.sum(PlatformLedgerEntry.amount), 0),
                func.count(),
            ).group_by(
                PlatformLedgerEntry.direction, PlatformLedgerEntry.source_type
            )
        )
    ).all()

    by_source: dict[str, Decimal] = {}
    total_credit = total_debit = Decimal("0.00")
    entries = 0
    for direction, source_type, amount, count in rows:
        entries += count
        if direction == LedgerDirection.CREDIT:
            total_credit += amount
            by_source[source_type.value] = (
                by_source.get(source_type.value, Decimal("0.00")) + amount
            )
        else:
            total_debit += amount

    return RevenueDashboard(
        currency=account.currency,
        wallet_balance=account.balance,
        commission_rate=account.commission_rate,
        total_commission_collected=total_credit,
        total_debited=total_debit,
        by_source=by_source,
        ledger_entries=entries,
    )


# ===========================================================================
# GET /dashboard/top-rooms
# ===========================================================================
@router.get("/dashboard/top-rooms", response_model=list[TopRoom])
async def top_rooms(ctx: AdminCtx, session: ScopedSession) -> list[TopRoom]:
    """Top 5 most-demanded rooms platform-wide (live bookings only)."""
    demand = func.count(Booking.id).label("demand")
    rows = (
        await session.execute(
            select(
                Room.id,
                Room.room_number,
                Tenant.name,
                demand,
                func.coalesce(func.sum(Booking.total_amount), 0).label("gross"),
            )
            .join(Booking, Booking.room_id == Room.id)
            .join(Tenant, Room.tenant_id == Tenant.id)
            .where(
                Booking.status.notin_(
                    [BookingStatus.CANCELLED, BookingStatus.NO_SHOW]
                )
            )
            .group_by(Room.id, Room.room_number, Tenant.name)
            .order_by(demand.desc(), func.sum(Booking.total_amount).desc())
            .limit(5)
        )
    ).all()
    return [
        TopRoom(
            room_id=room_id,
            room_number=room_number,
            hotel_name=hotel_name,
            demand=demand_count,
            gross_revenue=gross,
        )
        for room_id, room_number, hotel_name, demand_count, gross in rows
    ]


# ===========================================================================
# GET /export/revenue — styled .xlsx, streamed as a download
# ===========================================================================
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_MONEY_FORMAT = "#,##0.00"


def _style_header(worksheet, columns: list[str], widths: list[int]) -> None:
    for index, (title, width) in enumerate(zip(columns, widths), start=1):
        cell = worksheet.cell(row=1, column=index, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"


@router.get("/export/revenue")
async def export_revenue(ctx: AdminCtx, session: ScopedSession) -> StreamingResponse:
    """
    Excel workbook, two sheets:

    * **Monthly Revenue** — ledger credits per month per source, with a
      monthly net (credits − debits) column.
    * **Minibar Statistics** — per hotel per item: quantity consumed,
      gross revenue, settled vs outstanding.
    """
    month = func.to_char(
        func.date_trunc("month", PlatformLedgerEntry.created_at), "YYYY-MM"
    ).label("month")
    signed = func.sum(
        case(
            (
                PlatformLedgerEntry.direction == LedgerDirection.CREDIT,
                PlatformLedgerEntry.amount,
            ),
            else_=-PlatformLedgerEntry.amount,
        )
    )
    revenue_rows = (
        await session.execute(
            select(
                month,
                PlatformLedgerEntry.source_type,
                func.sum(PlatformLedgerEntry.amount),
                func.count(),
                signed,
            )
            .group_by(month, PlatformLedgerEntry.source_type)
            .order_by(month, PlatformLedgerEntry.source_type)
        )
    ).all()

    minibar_rows = (
        await session.execute(
            select(
                Tenant.name,
                MinibarConsumption.item_name,
                func.sum(MinibarConsumption.quantity),
                func.sum(
                    MinibarConsumption.unit_price * MinibarConsumption.quantity
                ),
                func.sum(
                    case((MinibarConsumption.is_settled.is_(True), 1), else_=0)
                ),
            )
            .join(Tenant, MinibarConsumption.tenant_id == Tenant.id)
            .group_by(Tenant.name, MinibarConsumption.item_name)
            .order_by(Tenant.name, MinibarConsumption.item_name)
        )
    ).all()

    workbook = Workbook()

    # -- Sheet 1: Monthly Revenue ----------------------------------------- #
    sheet = workbook.active
    sheet.title = "Monthly Revenue"
    _style_header(
        sheet,
        ["Month", "Source", "Amount (MNT)", "Entries", "Net Movement (MNT)"],
        [12, 28, 18, 10, 20],
    )
    row_index = 2
    grand_total = Decimal("0.00")
    for month_label, source_type, amount, count, net in revenue_rows:
        sheet.cell(row=row_index, column=1, value=month_label)
        sheet.cell(row=row_index, column=2, value=source_type.value)
        sheet.cell(row=row_index, column=3, value=float(amount)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=4, value=count)
        sheet.cell(row=row_index, column=5, value=float(net)).number_format = (
            _MONEY_FORMAT
        )
        grand_total += net
        row_index += 1
    total_label = sheet.cell(row=row_index, column=2, value="TOTAL NET")
    total_label.font = Font(bold=True)
    total_cell = sheet.cell(row=row_index, column=5, value=float(grand_total))
    total_cell.font = Font(bold=True)
    total_cell.number_format = _MONEY_FORMAT

    # -- Sheet 2: Minibar Statistics --------------------------------------- #
    sheet = workbook.create_sheet("Minibar Statistics")
    _style_header(
        sheet,
        ["Hotel", "Item", "Qty Consumed", "Revenue (MNT)", "Settled Lines"],
        [24, 28, 14, 16, 14],
    )
    for row_index, (hotel, item, qty, revenue, settled) in enumerate(
        minibar_rows, start=2
    ):
        sheet.cell(row=row_index, column=1, value=hotel)
        sheet.cell(row=row_index, column=2, value=item)
        sheet.cell(row=row_index, column=3, value=int(qty))
        sheet.cell(row=row_index, column=4, value=float(revenue)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=5, value=int(settled))

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = (
        f"platform_revenue_{datetime.now(timezone.utc):%Y%m%d_%H%M}.xlsx"
    )
    return StreamingResponse(
        buffer,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
